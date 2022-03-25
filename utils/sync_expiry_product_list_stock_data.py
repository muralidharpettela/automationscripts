import openpyxl as xl
import csv
import sys
# Source coming from shop
# Destination products in website sheet
# opening the source excel file
filename_kassen_system = r"/Users/muralidharpettela/Downloads/ks_dir/BK_Artikeldaten_25032022.csv"
#wb1 = xl.load_workbook(filename_kassen_system)
#ws1 = wb1.worksheets[0]

# opening the destination excel file
filename_website = r"/Users/muralidharpettela/Downloads/product_expiry/products_expiry_list.xlsx"
wb2 = xl.load_workbook(filename_website)
ws2 = wb2.worksheets[0]


def csv_to_excel(filename_kassen_system):
    if not ".csv" in filename_kassen_system:
        sys.stderr.write("Error: File does not have the ending \".csv\".\n")
        sys.exit(2)

    input_fh = open(filename_kassen_system, encoding="ISO-8859-1")
    workbook = xl.Workbook()
    # sheet = workbook.create_sheet(0)
    sheet = workbook.active

    for row_index, row in enumerate(
            csv.reader(open(filename_kassen_system, encoding="ISO-8859-1"), delimiter=";")):
        for col_index, col in enumerate(row):
            if row_index > 0:
                if col_index == 2 or col_index == 3 or col_index == 4 or col_index == 5:
                    sheet.cell(row=row_index + 1, column=col_index + 1).value = float(col.replace(",", "."))
                else:
                    sheet.cell(row=row_index + 1, column=col_index + 1).value = col
            else:
                sheet.cell(row=row_index + 1, column=col_index + 1).value = col

    # workbook.save(open(input_csv_file.replace(".csv", ".xlsx"), "wb"))
    return workbook

kassen_system_workbook = csv_to_excel(filename_kassen_system)
ws1 = kassen_system_workbook.worksheets[0]
# calculate total number of rows and
# columns in source excel file
mr_s = ws1.max_row
mc_s = ws1.max_column

mr_d = ws2.max_row
mc_d = ws2.max_column
# column of sort id source file
product_names_kassen_system = ws1['B'][1:]
# column of stock Source file
stock_kassen_system = ws1['F'][1:]
# column of stock Source file
price_kassen_system = ws1['C'][1:]
# Kassen system sale price (EK)
sale_price_kassen_system = ws1['D'][1:]
# tax class
tax_class_kassen_system = ws1['E'][1:]

# column of stock destination
product_names_website = ws2['B'][1:]
# column of stock destination
stock_website = ws2['D'][1:]
# column of price
#price_website = ws2['Z'][1:]
# sale price website
#sale_price_website = ws2['Y'][1:]
# tax class
#tax_class_website = ws2['M'][1:]

match_of_stock_cells_count = 0
num_of_product_stock_changed = 0
num_of_product_price_changed = 0
num_of_tax_class_changed = 0
num_no_match_found = 0
num_of_sale_price_updates = 0

no_match_products_list = list()
no_match_products_txt = open("../no_match_products.txt", "w+")
for i, product_website in enumerate(product_names_website):
    for j, product_kassen_system in enumerate(product_names_kassen_system):
        # check the sort id source and destination are same, if yes update the stock of destination with stock of source
        if str(product_website.value).rstrip() == str(product_kassen_system.value).rstrip():
            # stock update
            stock_website[i].value = stock_kassen_system[j].value
            match_of_stock_cells_count = match_of_stock_cells_count + 1
            break
        if (j == len(product_names_kassen_system) - 1):
            if str(product_website) not in no_match_products_list:
                no_match_products_list.append(product_website.value)
                no_match_products_txt.write(product_website.value)
                no_match_products_txt.write("\n")
                num_no_match_found = num_no_match_found + 1
print("Total no of Rows/Products in Source file from Shop File:{}".format(mr_s))
print("Total no of Rows/Products in Destination file in Website:{}".format(mr_d))
print("Number of Products Matched:{}".format(match_of_stock_cells_count))
print("Number of Products Stock Changed:{}".format(num_of_product_stock_changed))
print("Number of Products Price Changed:{}".format(num_of_product_price_changed))
print("Number of Products Tax Class Changed:{}".format(num_of_tax_class_changed))
print("Number of Products are no matched:{}".format(num_no_match_found))
print("Number of Products Sale Price Changed:{}".format(num_of_sale_price_updates))
# # copying the cell values from source
# # excel file to destination excel file
# for i in range(1, mr + 1):
#     for j in range(1, mc + 1):
#         # reading cell value from source excel file
#         c = ws1.cell(row=i, column=j)
#
#         # writing the read value to destination excel file
#         ws2.cell(row=i, column=j).value = c.value

# saving the destination excel file
wb2.save(str(filename_website))