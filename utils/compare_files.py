from common.common_functions import CommonFunctions
filepath_kassen_system = r"/Users/muralidharpettela/Downloads/BK_Artikeldaten_16042022.csv"
json_file_path = "../products.json"
no_match_products_list = list()
no_match_products_txt = open("../no_match_products.txt", "w+")

class CompareFiles(CommonFunctions):
    def __init__(self, filepath_kassen_system, json_file_path):
        super().__init__(filepath_kassen_system, json_file_path)
        from openpyxl import Workbook
        self.wb = Workbook()  # object of Workbook type
        print(self.wb.sheetnames)
        self.wb['Sheet'].title = "Products list"
        self.sh1 = self.wb.active  # Activate the sheet
        self.sh1['A1'].value = "Article"  # Writing into the cell
        self.sh1['B1'].value = "Price"
        self.sh1['C1'].value = "Sale Price"
        self.sh1['D1'].value = "Stock"


    def match_products(self, json_data_dict, kassen_system_data_dict):
        num_no_match_found = 0
        match_of_stock_cells_count = 0
        for i, product_kassen_system in enumerate(kassen_system_data_dict["product_names"]):
            for j, product_website in enumerate(json_data_dict):
                # check the sort id source and destination are same, if yes update the stock of destination with stock of source
                if str(product_website['name']).rstrip() == str(product_kassen_system.value).rstrip():

                    match_of_stock_cells_count = match_of_stock_cells_count + 1
                    break

                if (j == len(json_data_dict) - 1):
                    if str(str(product_kassen_system.value).rstrip()) not in no_match_products_list:
                        self.sh1.cell(row=num_no_match_found+2, column=1).value = kassen_system_data_dict["product_names"][i].value
                        self.sh1.cell(row=num_no_match_found+2, column=2).value = kassen_system_data_dict["price"][i].value
                        self.sh1.cell(row=num_no_match_found+2, column=3).value = kassen_system_data_dict["sale_price"][i].value
                        self.sh1.cell(row=num_no_match_found+2, column=4).value = kassen_system_data_dict["stock"][i].value
                        no_match_products_list.append(str(product_kassen_system.value).rstrip())
                        no_match_products_txt.write(str(product_kassen_system.value).rstrip())
                        no_match_products_txt.write("\n")
                        num_no_match_found = num_no_match_found + 1
        no_match_products_txt.close()
        self.wb.save("FirstCreatedPythonExcel1.xlsx")
        return match_of_stock_cells_count, num_no_match_found

    def main(self):
        workbook = self.csv_to_excel()
        ks_dict, mr_s, mc_s = self.load_kassen_system_excel_file(workbook)
        json_data = self.load_json_data_website_products()
        # kassen_system_data = self.assign_data_from_ks(ws1)
        match_of_stock_cells_count, num_no_match_found = self.match_products(json_data, ks_dict)
        print()


if __name__ == "__main__":
    filepath_kassen_system = r"/Users/muralidharpettela/Downloads/BK_Artikeldaten_16042022.csv"
    json_file_path = "./update_stock/products.json"
    cls_init = CompareFiles(filepath_kassen_system, json_file_path)
    cls_init.main()