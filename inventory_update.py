import openpyxl as xl
import math
# Source coming from shop
# Destination products in website sheet
# opening the source excel file
filename_kassen_system = r"/Users/muralidharpettela/Downloads/BK_Artikeldaten_03022022.xlsx"
wb1 = xl.load_workbook(filename_kassen_system)
ws1 = wb1.worksheets[0]

# opening the destination excel file
filename_website = r"/Users/muralidharpettela/Downloads/upload_products03022022.xlsx"
wb2 = xl.load_workbook(filename_website)
ws2 = wb2.worksheets[0]

# calculate total number of rows and
# columns in source excel file
mr_s = ws1.max_row
mc_s = ws1.max_column

mr_d = ws2.max_row
mc_d = ws2.max_column
# column of sort id source file
product_names_kassen_system = ws1['B'][1:]
# column of stock Source file
stock_kassen_system = ws1['F'][1:]
# column of stock Source file
price_kassen_system = ws1['C'][1:]
# Kassen system sale price (EK)
sale_price_kassen_system = ws1['D'][1:]
# tax class
tax_class_kassen_system = ws1['E'][1:]

# column of stock destination
product_names_website = ws2['D'][1:]
# column of stock destination
stock_website = ws2['O'][1:]
# column of price
price_website = ws2['Z'][1:]
# sale price website
sale_price_website = ws2['Y'][1:]
# tax class
tax_class_website = ws2['M'][1:]

match_of_stock_cells_count = 0
num_of_product_stock_changed = 0
num_of_product_price_changed = 0
num_of_tax_class_changed = 0
num_no_match_found = 0
num_of_sale_price_updates = 0

no_match_products_list = list()
no_match_products_txt = open("no_match_products.txt", "w+")
for i, product_website in enumerate(product_names_website):
    for j, product_kassen_system in enumerate(product_names_kassen_system):
        # check the sort id source and destination are same, if yes update the stock of destination with stock of source
        if str(product_website.value).rstrip() == str(product_kassen_system.value).rstrip():
            # weight calculate
            if str(product_website.value).rstrip().count("-") > 1:
                # index_of_ = str(product_website.value).rstrip().index("-")
                # new_product_name = str(product_website.value)[index_of_].replace("-", " ")
                pass
            else:
                if len(str(product_website.value).rstrip().split("-")[1].split()) > 2:
                    weight_grams_list = str(product_website.value).rstrip().split("-")[1].split()[:-1]
                else:
                    weight_grams_list = str(product_website.value).rstrip().split("-")[1].split()
                try:
                    temp = i + 1
                    if weight_grams_list[1] == "g":
                        weight = float(weight_grams_list[0]) / 1000
                        # weight
                        ws2['S'][temp].value = str(weight).replace(",", ".")
                    elif weight_grams_list[1] == "kg":
                        weight = float(weight_grams_list[0])
                        # weight
                        ws2['S'][temp].value = str(weight).replace(",", ".")
                    elif weight_grams_list[1] == "l":
                        weight = float(weight_grams_list[0])
                        # weight
                        ws2['S'][temp].value = str(weight).replace(",", ".")
                    elif weight_grams_list[1] == "ml":
                        weight = float(weight_grams_list[0]) / 1000
                        # weight
                        ws2['S'][temp].value = str(weight).replace(",", ".")
                    else:
                        pass
                except:
                    print(str(product_website.value).rstrip())
            if stock_website[i].value != stock_kassen_system[j].value:
                num_of_product_stock_changed = num_of_product_stock_changed + 1
            # price_str = str(price_kassen_system[j].value).replace(",", ".")
            # tax_amount = float(price_str) * (int(tax_class_kassen_system[j].value) / 100)
            # price_including_tax = round(float(price_website[i].value) + tax_amount, 2)
            # if price_including_tax.is_integer() == True:
            #     price_including_tax = str(price_including_tax)[:-2]
            #     pass
            # else:
            #     price_including_tax = str(price_including_tax).replace(",", ".")
            if price_website[i].value != str(price_kassen_system[j].value).replace(",", "."):
                num_of_product_price_changed = num_of_product_price_changed + 1
            if sale_price_website[i].value != str(sale_price_kassen_system[j].value).replace(",", "."):
                num_of_sale_price_updates = num_of_sale_price_updates + 1
            if tax_class_website[i].value != tax_class_kassen_system[j].value:
                num_of_tax_class_changed = num_of_tax_class_changed + 1
            # stock update
            stock_website[i].value = stock_kassen_system[j].value
            # price update
            # remove the tax rate from price
            # price_str = str(price_kassen_system[j].value).replace(",", ".")
            # tax_amount = float(price_str) * (int(tax_class_kassen_system[j].value)/100)
            # price_excluding_tax = round(float(price_str) - tax_amount, 2)
            # price_excluding_tax = str(price_excluding_tax).replace(",", ".")
            # price_website[i].value = price_excluding_tax
            if sale_price_kassen_system[j].value == 0:
                price_website[i].value = str(price_kassen_system[j].value).replace(",", ".")
                #sale_price_website[i].value = str(sale_price_kassen_system[j].value).replace(",", ".")
            else:
                price_website[i].value = str(sale_price_kassen_system[j].value).replace(",", ".")
                sale_price_website[i].value = str(price_kassen_system[j].value).replace(",", ".")
            # tax class update
            if tax_class_kassen_system[j].value == 7:
                tax_class_website[i].value = "Tax 7 Per"
            else:
                tax_class_website[i].value = "Tax 19 Per"
            match_of_stock_cells_count = match_of_stock_cells_count + 1
            break
        if (j == len(product_names_kassen_system) - 1):
            if str(product_website) not in no_match_products_list:
                no_match_products_list.append(product_website.value)
                no_match_products_txt.write(product_website.value)
                no_match_products_txt.write("\n")
                num_no_match_found = num_no_match_found + 1
print("Total no of Rows/Products in Source file from Shop File:{}".format(mr_s))
print("Total no of Rows/Products in Destination file in Website:{}".format(mr_d))
print("Number of Products Matched:{}".format(match_of_stock_cells_count))
print("Number of Products Stock Changed:{}".format(num_of_product_stock_changed))
print("Number of Products Price Changed:{}".format(num_of_product_price_changed))
print("Number of Products Tax Class Changed:{}".format(num_of_tax_class_changed))
print("Number of Products are no matched:{}".format(num_no_match_found))
print("Number of Products Sale Price Changed:{}".format(num_of_sale_price_updates))
# # copying the cell values from source
# # excel file to destination excel file
# for i in range(1, mr + 1):
#     for j in range(1, mc + 1):
#         # reading cell value from source excel file
#         c = ws1.cell(row=i, column=j)
#
#         # writing the read value to destination excel file
#         ws2.cell(row=i, column=j).value = c.value

# saving the destination excel file
wb2.save(str(filename_website))