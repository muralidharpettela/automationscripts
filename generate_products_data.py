import openpyxl as xl
from os import listdir
from os.path import isfile, join
images_path = r"/Users/muralidharpettela/Downloads/"
# Source coming from shop
# Destination products in website sheet
# opening the source excel file
filename_excel_path = r"/Users/muralidharpettela/Downloads/Book1.xlsx"
wb1 = xl.load_workbook(filename_excel_path)
ws1 = wb1.worksheets[0]

# calculate total number of rows and
# columns in source excel file
mr_s = ws1.max_row
mc_s = ws1.max_column

# column of sort id source file
col_name = ws1['D'][1:]
# category data
col_category = ws1['AA'][1:]


onlyfiles = [f for f in listdir(images_path) if isfile(join(images_path, f))]

for i, name in enumerate(col_name):
    list_of_words = str(name.value).split()
    for j, word in enumerate(list_of_words):
        if word == "-":
            del list_of_words[j]
    final_string = "_".join(list_of_words)  # SKU generate
    # weight calculate
    if len(str(name.value).rstrip().split("-")[1].split()) > 2:
        weight_grams_list = str(name.value).rstrip().split("-")[1].split()[:-1]
    else:
        weight_grams_list = str(name.value).rstrip().split("-")[1].split()
    try:
        temp_1 = i + 1
        if weight_grams_list[1] == "g":
            weight = float(weight_grams_list[0]) / 1000
            # weight
            ws1['S'][temp_1].value = str(weight).replace(",", ".")
        elif weight_grams_list[1] == "kg":
            weight = float(weight_grams_list[0])
            # weight
            ws1['S'][temp_1].value = str(weight).replace(",", ".")
        elif weight_grams_list[1] == "l":
            weight = float(weight_grams_list[0])
            # weight
            ws1['S'][temp_1].value = str(weight).replace(",", ".")
        elif weight_grams_list[1] == "ml":
            weight = float(weight_grams_list[0]) / 1000
            # weight
            ws1['S'][temp_1].value = str(weight).replace(",", ".")
        else:
            pass
    except:
        print(str(name.value).rstrip())
    # weight_grams_list = str(name.value).split("-")[1].split()[:-1]
    # if weight_grams_list[1] == "g":
    #     weight = int(weight_grams_list[0])
    # elif weight_grams_list[1] == "kg":
    #     weight = int(weight_grams_list[0]) * 1000
    # elif weight_grams_list[1] == "l":
    #     weight = int(weight_grams_list[0]) * 1000
    # elif weight_grams_list[1] == "ml":
    #     weight = int(weight_grams_list[0])
    # else:
    #     weight = "None"

    temp = i + 1
    ws1['C'][temp].value = final_string  # assign SKU
    # assign type
    ws1['B'][temp].value = "simple"
    # published
    ws1['E'][temp].value = 1
    # Is Featured
    ws1['F'][temp].value = 0
    # visibility
    ws1['G'][temp].value = "visible"
    # tax status
    ws1['L'][temp].value = "taxable"
    # instock
    ws1['N'][temp].value = 0
    # Backorder
    ws1['Q'][temp].value = 0
    # sold individually
    ws1['R'][temp].value = 0
    # weight
    #ws1['S'][temp].value = str(weight).replace(",", ".")
    # sold individually
    ws1['W'][temp].value = 0
    # image
    if (name.value + ".jpg") in onlyfiles:
        print("Match:{}".format(name.value))
        index = onlyfiles.index(name.value + ".jpg")
        del onlyfiles[index]
        ws1['AD'][temp].value = str("https://www.lotus-grocery.eu/wp-content/uploads/2021/12/") + str(name.value) + ".jpg"
    # position
    ws1['AM'][temp].value = 0
    # category
    if col_category[i].value == "RI":
        ws1['AA'][temp].value = "Rice & Atta > Rice"
    elif col_category[i].value == "AT":
        ws1['AA'][temp].value = "Rice & Atta > Atta"
    elif col_category[i].value == "FL":
        ws1['AA'][temp].value = "Rava & Flour > Flour"
    elif col_category[i].value == "RA":
        ws1['AA'][temp].value = "Rava & Flour >  Rava"
    elif col_category[i].value == "BS":
        ws1['AA'][temp].value = "Masala & Spice Mix > Biryani Spices"
    elif col_category[i].value == "PS":
        ws1['AA'][temp].value = "Masala & Spice Mix > Powdered Spices"
    elif col_category[i].value == "SM":
        ws1['AA'][temp].value = "Masala & Spice Mix > Spice Mix"
    elif col_category[i].value == "WS":
        ws1['AA'][temp].value = "Masala & Spice Mix > Whole Spices"
    elif col_category[i].value == "TCB":
        ws1['AA'][temp].value = "Cooking Essentials & Beverages > Tea Coffee & Beverages"
    elif col_category[i].value == "CE":
        ws1['AA'][temp].value = "Cooking Essentials & Beverages > Cooking Essentials"
    elif col_category[i].value == "OG":
        ws1['AA'][temp].value = "Cooking Essentials & Beverages > Oils & Ghee"
    elif col_category[i].value == "GM":
        ws1['AA'][temp].value = "Food Grains & Pulses > Grains & Millets"
    elif col_category[i].value == "PL":
        ws1['AA'][temp].value = "Food Grains & Pulses > Pulses/Lentils"
    elif col_category[i].value == "PIC":
        ws1['AA'][temp].value = "Pickles Pastes Pulps & Sauces > Pickles"
    elif col_category[i].value == "PAS":
        ws1['AA'][temp].value = "Pickles Pastes Pulps & Sauces > Pastes"
    elif col_category[i].value == "PSY":
        ws1['AA'][temp].value = "Pickles Pastes Pulps & Sauces > Pulps & Syrups"
    elif col_category[i].value == "SAU":
        ws1['AA'][temp].value = "Pickles Pastes Pulps & Sauces > Sauces"
    elif col_category[i].value == "CF":
        ws1['AA'][temp].value = "Snacks > Chips & Fryums"
    elif col_category[i].value == "CBC":
        ws1['AA'][temp].value = "Snacks > Cake Biscuit & Chikkis"
    elif col_category[i].value == "DFS":
        ws1['AA'][temp].value = "Snacks >  Dry Fruits & Seeds"
    elif col_category[i].value == "SN":
        ws1['AA'][temp].value = "Snacks > Snacks & Namkeen"
    elif col_category[i].value == "SW":
        ws1['AA'][temp].value = "Snacks > Sweets"
    elif col_category[i].value == "PA":
        ws1['AA'][temp].value = "Breads & Pappadams > Papadam"
    elif col_category[i].value == "BC":
        ws1['AA'][temp].value = "Beauty & Hygiene Care > Body Care"
    elif col_category[i].value == "HC":
        ws1['AA'][temp].value = "Beauty & Hygiene Care > Hair Care"
    elif col_category[i].value == "HCO":
        ws1['AA'][temp].value = "Beauty & Hygiene Care > Health Care & Others"
    elif col_category[i].value == "IM":
        ws1['AA'][temp].value = "Ready Made Food products > Instant Mix"
    elif col_category[i].value == "RE":
        ws1['AA'][temp].value = "Ready Made Food products > Ready to Eat"
    elif col_category[i].value == "NO":
        ws1['AA'][temp].value = "Ready Made Food products > Noodles"
    elif col_category[i].value == "ID":
        ws1['AA'][temp].value = "Idols & Pooja Items > Idols"
    elif col_category[i].value == "PI":
        ws1['AA'][temp].value = "Idols & Pooja Items > Pooja Items"
    elif col_category[i].value == "PM":
        ws1['AA'][temp].value = "Idols & Pooja Items > Pooja Mandir"
    elif col_category[i].value == "KW":
        ws1['AA'][temp].value = "Kitchenware & Household Items > Kitchenware"
    elif col_category[i].value == "HI":
        ws1['AA'][temp].value = "Kitchenware & Household Items > Household Items"
    else:
        ws1['AA'][temp].value = "None"




print(onlyfiles)
    #saving the destination excel file
wb1.save(str(filename_excel_path))