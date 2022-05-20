from common.common_functions import CommonFunctions
from tabulate import tabulate
from openpyxl import Workbook


class NotifyNewProducts(CommonFunctions):
    def __init__(self, filepath_kassen_system, json_file_path, credentials_path):
        super().__init__(filepath_kassen_system, json_file_path)
        self.email_credentials = self.load_wp_credentials(credentials_path)
        with open("./notification_new_products/frozen_not_required_products.txt") as txt:
            self.frozen_not_rquired_products = txt.read().split("\n")
            self.frozen_not_rquired_products = [x.rstrip() for x in self.frozen_not_rquired_products]

        self.text = """
        Hello, Muralidhar

        You have a work to do, your script filtered a products that are not in website, it seems to be new products are added in kassensystem . Have a look and upload products as early as possible

        {table}
        {stat}

        Regards,

        Your Wonderful Script"""

        self.html = """
        <html><body><p>Hello, Muralidhar.</p>
        <p> You have a work to do, your script filtered a products that are not in website, it seems to be new products are added in kassensystem . Have a look and upload products as early as possible</p>
        {table}
        {stat}
        <p>Regards,</p>
        <p>Your Wonderful Script</p>
        </body></html>
        """
        self.wb = Workbook()  # object of Workbook type
        print(self.wb.sheetnames)
        self.wb['Sheet'].title = "Products list"
        self.sh1 = self.wb.active  # Activate the sheet
        self.sh1['A1'].value = "Article"  # Writing into the cell
        self.sh1['B1'].value = "Price"
        self.sh1['C1'].value = "Sale Price"
        self.sh1['D1'].value = "Stock"


    def match_products(self, json_data_dict, kassen_system_data_dict):
        num_no_match_found = 0
        match_of_stock_cells_count = 0
        for i, product_kassen_system in enumerate(kassen_system_data_dict["product_names"]):
            if not product_kassen_system.value.lstrip().rstrip() in self.frozen_not_rquired_products:
                for j, product_website in enumerate(json_data_dict):
                    # check the sort id source and destination are same, if yes update the stock of destination with stock of source
                    if str(product_website['name']).rstrip() == str(product_kassen_system.value).rstrip():
                        match_of_stock_cells_count = match_of_stock_cells_count + 1
                        break

                    if (j == len(json_data_dict) - 1):
                        if str(str(product_kassen_system.value).rstrip()) not in self.no_match_products_list and kassen_system_data_dict["stock"][i].value > 0:
                            self.sh1.cell(row=num_no_match_found+2, column=1).value = kassen_system_data_dict["product_names"][i].value
                            self.sh1.cell(row=num_no_match_found+2, column=2).value = kassen_system_data_dict["price"][i].value
                            self.sh1.cell(row=num_no_match_found+2, column=3).value = kassen_system_data_dict["sale_price"][i].value
                            self.sh1.cell(row=num_no_match_found+2, column=4).value = kassen_system_data_dict["stock"][i].value
                            self.no_match_products_list.append(str(product_kassen_system.value).rstrip())
                            self.no_match_products_txt.write(str(product_kassen_system.value).rstrip())
                            self.no_match_products_txt.write("\n")
                            num_no_match_found = num_no_match_found + 1
        self.no_match_products_txt.close()
        # self.wb.save("new_products_to_upload.csv")
        text = self.text.format(table=tabulate(self.wb.worksheets[0].values, headers="firstrow", tablefmt="grid"), stat=tabulate([['Products Matched', match_of_stock_cells_count], ['Products Not Matched', num_no_match_found]], headers=['Stat Name', 'Stat Value'],tablefmt="grid"))
        html = self.html.format(table=tabulate(self.wb.worksheets[0].values, headers="firstrow", tablefmt="html"), stat=tabulate([['Products Matched', match_of_stock_cells_count], ['Products Not Matched', num_no_match_found]], headers=['Stat Name', 'Stat Value'],tablefmt="html"))
        subject = "[Notification] - lotus-grocery.eu new Products detected that need to be add in Website"
        self.send_email_using_smtp(text, html, subject, "stocks@lotus-grocery.eu", "info@lotus-grocery.eu")
        return match_of_stock_cells_count, num_no_match_found

    def main(self):
        workbook = self.csv_to_excel()
        ks_dict, mr_s, mc_s = self.load_kassen_system_excel_file(workbook)
        json_data = self.load_json_data_website_products()
        # kassen_system_data = self.assign_data_from_ks(ws1)
        self.match_products(json_data, ks_dict)
        print()


if __name__ == "__main__":
    filepath_kassen_system = r"/Users/muralidharpettela/Downloads/ks_dir/BK_Artikeldaten_20052022_1.csv"
    json_file_path = "./update_stock/products.json"
    credentials_path = "./common/email_credentials.json"
    cls_init = NotifyNewProducts(filepath_kassen_system, json_file_path, credentials_path)
    cls_init.main()