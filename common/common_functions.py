import re
import json
import csv
import sys
from openpyxl.workbook import Workbook
import yagmail
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import smtplib

class CommonFunctions:
    def __init__(self, filepath_kassen_system, json_file_path="products.json"):
        self.filepath_kassen_system = filepath_kassen_system
        self.json_file_path = json_file_path

        self.products_list = list()
        self.no_match_products_list = list()
        self.no_match_products_txt = open("./no_match_products.txt", "w+")
        self.products_without_weight_txt = open("./products_without_weight.txt", "w+")

    def load_wp_credentials(self, json_file_path):
        # load the json file
        # Opening JSON file
        f = open(json_file_path)
        # returns JSON object as
        # a dictionary
        data = json.load(f)
        return data

    def csv_to_excel(self):
        if not ".csv" in self.filepath_kassen_system:
            sys.stderr.write("Error: File does not have the ending \".csv\".\n")
            sys.exit(2)

        input_fh = open(self.filepath_kassen_system, encoding="ISO-8859-1")
        workbook = Workbook()
        # sheet = workbook.create_sheet(0)
        sheet = workbook.active

        for row_index, row in enumerate(csv.reader(open(self.filepath_kassen_system, encoding="ISO-8859-1"), delimiter=";")):
            for col_index, col in enumerate(row):
                if row_index > 0:
                    if col_index == 2 or col_index == 3 or col_index == 4 or col_index == 5:
                        sheet.cell(row=row_index + 1, column=col_index + 1).value = float(col.replace(",", "."))
                    else:
                        sheet.cell(row=row_index + 1, column=col_index + 1).value = col
                else:
                    sheet.cell(row=row_index + 1, column=col_index + 1).value = col

        # workbook.save(open(input_csv_file.replace(".csv", ".xlsx"), "wb"))
        return workbook

    def load_kassen_system_excel_file(self, workbook):
        # wb1 = xl.load_workbook(filename_kassen_system_path)
        kassen_system_data_dict = {"product_names": list, "stock": list, "price": list, "sale_price": list,
                                        "tax_class": list}
        ws1 = workbook.worksheets[0]
        # calculate total number of rows and
        # columns in source excel file
        mr_s = ws1.max_row
        mc_s = ws1.max_column
        # column of sort id source file
        kassen_system_data_dict["product_names"] = ws1['B'][1:]
        # column of stock Source file
        kassen_system_data_dict["stock"] = ws1['F'][1:]
        # column of stock Source file
        kassen_system_data_dict["price"] = ws1['C'][1:]
        # Kassen system sale price (EK)
        kassen_system_data_dict["sale_price"] = ws1['D'][1:]
        # tax class
        kassen_system_data_dict["tax_class"] = ws1['E'][1:]

        return kassen_system_data_dict, mr_s, mc_s

    def load_kassen_system_expiry_excel_file(self, workbook):
        # wb1 = xl.load_workbook(filename_kassen_system_path)
        products_expiry_data_dict = {"product_names": list, "stock": list, "expiry_date":list}
        ws1 = workbook.worksheets[0]
        # calculate total number of rows and
        # columns in source excel file
        mr_s = ws1.max_row
        mc_s = ws1.max_column
        # column of sort id source file
        products_expiry_data_dict["product_names"] = ws1['B'][1:]
        products_expiry_data_dict["stock"] = ws1['D'][1:]
        products_expiry_data_dict["expiry_date"] = ws1['E'][1:]

        return products_expiry_data_dict, mr_s, mc_s

    def load_json_data_website_products(self):
        # load the json file
        # Opening JSON file
        f = open(self.json_file_path)
        # returns JSON object as
        # a dictionary
        data = json.load(f)
        return data

    def calculate_weight(self, product_website, products_kassen_system_dict):
        # weight calculate
        try:
            weight_value_with_unit = re.search(r'([0-9]+[" "]+(g|ml|kg|l))', str(product_website['name'])).group(1)
            split_weight_unit = weight_value_with_unit.split(" ")
            try:
                if split_weight_unit[1] == "g":
                    weight = float(split_weight_unit[0]) / 1000
                    # weight
                    products_kassen_system_dict['weight'] = str(weight).replace(",", ".")
                elif split_weight_unit[1] == "kg":
                    weight = float(split_weight_unit[0])
                    # weight
                    products_kassen_system_dict['weight'] = str(weight).replace(",", ".")
                elif split_weight_unit[1] == "l":
                    weight = float(split_weight_unit[0])
                    # weight
                    products_kassen_system_dict['weight'] = str(weight).replace(",", ".")
                elif split_weight_unit[1] == "ml":
                    weight = float(split_weight_unit[0]) / 1000
                    # weight
                    products_kassen_system_dict['weight'] = str(weight).replace(",", ".")
                else:
                    pass
            except:
                print(str(product_website['name']).rstrip())
        except:
            self.products_without_weight_txt.write(str(product_website['name']))
            self.products_without_weight_txt.write("\n")

    def match_products_and_update(self, json_data_dict, kassen_system_data_dict):
        num_no_match_found = 0
        match_of_stock_cells_count = 0
        weight_updated_products = 0

        for i, product_website in enumerate(json_data_dict):
            products_kassen_system_dict = {"id": 0, "weight": "0", "stock_quantity": 0, "regular_price": 0,
                                           "sale_price": None,
                                           "tax_class": None}
            for j, product_kassen_system in enumerate(kassen_system_data_dict["product_names"]):
                # check the sort id source and destination are same, if yes update the stock of destination with stock of source
                if str(product_website['name']).rstrip() == str(product_kassen_system.value).rstrip():
                    # product brand attribute
                    # result = re.search(r'^([a-zA-Z_\s\-]+[" "])', str(product_website['name']).rstrip())
                    # if result:
                    #     products_kassen_system_dict = {"id": 0, "weight": "0", "stock_quantity": 0, "regular_price": 0,
                    #                                    "sale_price": None,
                    #                                    "tax_class": None, 'attributes': [
                    #             {'id': 3, 'name': 'Brand', 'position': 0, 'visible': True, 'variation': False,
                    #              'options': None}]
                    #                                    }
                    #     products_kassen_system_dict['attributes'][0]['options'] = [result.group(1)[:-3]]
                    # else:
                    #     products_kassen_system_dict = {"id": 0, "weight": "0", "stock_quantity": 0, "regular_price": 0,
                    #                                    "sale_price": None,
                    #                                    "tax_class": None}
                    #     #print(result.group(1)[:-3])
                    products_kassen_system_dict['id'] = product_website["id"]
                    # calculate weight
                    #if float(product_website['weight']) == 0:
                        #weight_updated_products = weight_updated_products + 1
                    self.calculate_weight(product_website, products_kassen_system_dict)
                    # stock update
                    products_kassen_system_dict['stock_quantity'] = kassen_system_data_dict["stock"][j].value
                    # price update
                    # products_kassen_system_dict['price'] = str(kassen_system_data_dict["price"][j].value)
                    # sale price
                    if kassen_system_data_dict["sale_price"][j].value != 0:
                        products_kassen_system_dict['sale_price'] = str(kassen_system_data_dict["price"][j].value)
                        products_kassen_system_dict['regular_price'] = str(
                            kassen_system_data_dict["sale_price"][j].value)
                    else:
                        # regular price update
                        products_kassen_system_dict['regular_price'] = str(kassen_system_data_dict["price"][j].value)
                    # tax class update
                    if kassen_system_data_dict["tax_class"][j].value == 7:
                        products_kassen_system_dict['tax_class'] = "Tax 7 Per"
                    else:
                        products_kassen_system_dict['tax_class'] = "Tax 19 Per"

                    # print(wcapi.put("products/" + str(product_website["id"]), products_kassen_system_dict).json())
                    # wcapi.put("products/" + str(product_website["id"]), products_kassen_system_dict).json()
                    self.products_list.append(products_kassen_system_dict)
                    match_of_stock_cells_count = match_of_stock_cells_count + 1
                    break

                if (j == len(kassen_system_data_dict["product_names"]) - 1):
                    if str(product_website['name']) not in self.no_match_products_list:
                        self.no_match_products_list.append(product_website['name'])
                        self.no_match_products_txt.write(product_website['name'])
                        self.no_match_products_txt.write("\n")
                        num_no_match_found = num_no_match_found + 1
        self.no_match_products_txt.close()
        self.products_without_weight_txt.close()
        return weight_updated_products, match_of_stock_cells_count, num_no_match_found

    def load_json_data_website_products_mhd(self, json_file_path):
        # load the json file
        # Opening JSON file
        f = open(json_file_path)
        # returns JSON object as
        # a dictionary
        data = json.load(f)
        return data

    def match_products_and_update_mhd(self, json_data_dict, kassen_system_data_dict):
        num_no_match_found = 0
        match_of_stock_cells_count = 0
        weight_updated_products = 0

        for i, product_website in enumerate(json_data_dict):
            for j, product_kassen_system in enumerate(kassen_system_data_dict["product_names"]):
                # check the sort id source and destination are same, if yes update the stock of destination with stock of source
                if str(product_website['name']).rstrip() == str(product_kassen_system.value).rstrip():
                    products_kassen_system_dict = {"id": 0, 'attributes': [
                        {'id': 3, 'name': 'Brand', 'position': 0, 'visible': True, 'variation': False, 'options': None},
                        {"id": 5, "name": "MHD", "position": 1, "visible": True, "variation": False, "options": None}]}
                    result = re.search(r'^([a-zA-Z_\s\-]+[" "])', str(product_website['name']).rstrip())
                    if kassen_system_data_dict['stock'][j].value == 0:
                        # check for existence of MHD attribute
                        try:
                            if product_website['attributes'][1]:
                                # make to None, or delete MHD attribute
                                if product_website['attributes'][1]['id'] == 5:
                                    products_kassen_system_dict['id'] = product_website["id"]
                                    if result:
                                        products_kassen_system_dict['attributes'][0]['options'] = [result.group(1)[:-3]]
                                    self.products_list.append(products_kassen_system_dict)
                        except IndexError:
                            print("Sorry, Index of MHD not available")
                    if kassen_system_data_dict['expiry_date'][j].value:
                        if kassen_system_data_dict['stock'][j].value > 0 and kassen_system_data_dict['expiry_date'][j].value.date() <= (date.today() + relativedelta(months=+3)):
                            products_kassen_system_dict['attributes'][1]["options"] = ["MHD: " + kassen_system_data_dict['expiry_date'][j].value.date().strftime("%d-%m-%Y")]
                            products_kassen_system_dict['id'] = product_website["id"]
                            if result:
                                products_kassen_system_dict['attributes'][0]['options'] = [result.group(1)[:-3]]
                            self.products_list.append(products_kassen_system_dict)
                        if kassen_system_data_dict['stock'][j].value > 0 and kassen_system_data_dict['expiry_date'][j].value.date() >= (date.today() + relativedelta(months=+3)):
                            # check for existence of MHD attribute
                            # check for existence of MHD attribute
                            try:
                                if product_website['attributes'][1]:
                                    # make to None, or delete MHD attribute
                                    if product_website['attributes'][1]['id'] == 5:
                                        products_kassen_system_dict['id'] = product_website["id"]
                                        if result:
                                            products_kassen_system_dict['attributes'][0]['options'] = [
                                                result.group(1)[:-3]]
                                        self.products_list.append(products_kassen_system_dict)
                            except IndexError:
                                print("Sorry, Index of MHD not available")
                    match_of_stock_cells_count = match_of_stock_cells_count + 1
                    break

                if (j == len(kassen_system_data_dict["product_names"]) - 1):
                    if str(product_website['name']) not in self.no_match_products_list:
                        self.no_match_products_list.append(product_website['name'])
                        self.no_match_products_txt.write(product_website['name'])
                        self.no_match_products_txt.write("\n")
                        num_no_match_found = num_no_match_found + 1
        self.no_match_products_txt.close()
        self.products_without_weight_txt.close()
        return weight_updated_products, match_of_stock_cells_count, num_no_match_found

    def send_email(self, subject, content):
        user = 'lotusgroceryingolstadt2@gmail.com'
        app_password = 'ofanobixbzpvcpqz'  # a token for gmail
        to = 'info@lotus-grocery.eu'  # To send a group of recipients, simply change ‘to’ to a list.

        subject = subject
        content = content

        with yagmail.SMTP(user, app_password) as yag:
            yag.send(to, subject, content)
            print('Sent email successfully')

    def send_email_using_smtp(self, text, html, subject, from_mail, to_mail):
        message = MIMEMultipart(
            "alternative", None, [MIMEText(text), MIMEText(html, 'html')])

        message['Subject'] = subject
        message['From'] = from_mail
        message['To'] = to_mail
        server = smtplib.SMTP(self.email_credentials["server"])
        server.ehlo()
        server.starttls()
        server.login(self.email_credentials["username"], self.email_credentials["password"])
        server.sendmail(from_mail, to_mail, message.as_string())
        server.quit()
