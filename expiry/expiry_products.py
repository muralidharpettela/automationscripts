import openpyxl as xl
import base64
import glob
import os
import requests
from datetime import datetime
from datetime import date
from dateutil.relativedelta import relativedelta
from common.common_functions import CommonFunctions
from tabulate import tabulate

class ExpiryProducts(CommonFunctions):
    def __init__(self, kassen_system_excel_file_dir, product_expiry_list_dir, onedrive_link, credentials_path):
        self.kassen_system_excel_file_dir = kassen_system_excel_file_dir
        self.product_expiry_list_dir = product_expiry_list_dir
        # onedrive file link of products expiry list
        self.onedrive_link = onedrive_link
        # dates declare for next 3 months
        self.one_months = date.today() + relativedelta(months=+1)
        self.two_months = date.today() + relativedelta(months=+2)
        self.three_months = date.today() + relativedelta(months=+3)
        # declare lists greater than 10
        self.ten_one_month_expiry_list = list()
        self.ten_two_month_expiry_list = list()
        self.ten_three_month_expiry_list = list()
        # declare lists greater than 20
        self.twenty_one_month_expiry_list = list()
        self.twenty_two_month_expiry_list = list()
        self.twenty_three_month_expiry_list = list()
        # declare list for below 10
        self.below_ten_one_month_expiry_list = list()
        self.below_ten_two_month_expiry_list = list()
        self.below_ten_three_month_expiry_list = list()
        self.products_already_expired_list = list()
        self.no_expiry_date_products = list()
        self.stock_zero_expiry_date_exists = list()
        self.stock_exist_expiry_date_not_exists = list()
        # email script
        self.text = """
               Hello, Lotus Grocery.
               This is an automated mail, notification of the expiry products within 3 months.
               {stock201month}
               {stock111month}
               {stock11month}
               {stock202month}
               {stock112month}
               {stock12month}
               {stock203month}
               {stock113month}
               {stock13month}
               {expiredproducts}
               {case1}
               {case2}
               {stat}
               
               *Case-1: Stock value 0 but the expiry date exists
               *Case-2: Stock value is not zero but the expiry date doesnot exists
               Regards,

               Your Expiry Product Script"""

        self.html = """
               <html><body><p>Hello, Lotus Grocery.</p>
               <p> This is an automated mail, notification of the expiry products within 3 months.</p>
               {stock201month}
               {stock111month}
               {stock11month}
               {stock202month}
               {stock112month}
               {stock12month}
               {stock203month}
               {stock113month}
               {stock13month}
               {expiredproducts}
               {case1}
               {case2}
               {stat}
               *Case-1: Stock value 0 but the expiry date exists
               *Case-2: Stock value is not zero but the expiry date doesnot exists
               <p>Regards,</p>
               <p>Your Expiry Product Script</p>
               </body></html>
               """
        self.email_credentials = self.load_wp_credentials(credentials_path)

    def create_onedrive_directdownload(self):
        data_bytes64 = base64.b64encode(bytes(self.onedrive_link, 'utf-8'))
        data_bytes64_String = data_bytes64.decode('utf-8').replace('/', '_').replace('+', '-').rstrip("=")
        resultUrl = f"https://api.onedrive.com/v1.0/shares/u!{data_bytes64_String}/root/content"
        return resultUrl

    def download_onedrive_file(self):
        onedrive_direct_link = self.create_onedrive_directdownload()
        r = requests.get(onedrive_direct_link)
        save_path = os.path.join(product_expiry_list_dir, "products_expiry_list.xlsx")
        with open(save_path, 'wb') as f:
            f.write(r.content)
        return save_path

    def get_kassen_system_latest_file(self):
        # kassen system latest file
        list_of_files = glob.glob(
            self.kassen_system_excel_file_dir + "/*.csv")  # * means all if need specific format then *.csv
        latest_file = max(list_of_files, key=os.path.getctime)
        print(latest_file)
        return latest_file

    def sync_stock_from_ks_to_expiry_list(self, product_expiry_dict, kassen_system_data_dict):
        match_of_stock_cells_count = 0
        num_no_match_found = 0

        no_match_products_list = list()
        no_match_products_txt = open("../no_match_products.txt", "w+")
        for i, product_name_expiry_list in enumerate(product_expiry_dict['product_names']):
            for j, product_kassen_system in enumerate(kassen_system_data_dict['product_names']):
                # check the sort id source and destination are same, if yes update the stock of destination with stock of source
                if str(product_name_expiry_list.value).rstrip() == str(product_kassen_system.value).rstrip():
                    # stock update
                    product_expiry_dict['stock'][i].value = kassen_system_data_dict['stock'][j].value
                    match_of_stock_cells_count = match_of_stock_cells_count + 1
                    break
                if (j == len(kassen_system_data_dict['product_names']) - 1):
                    if str(product_name_expiry_list.value) not in no_match_products_list:
                        no_match_products_list.append(product_name_expiry_list.value)
                        no_match_products_txt.write(product_name_expiry_list.value)
                        no_match_products_txt.write("\n")
                        num_no_match_found = num_no_match_found + 1

        return match_of_stock_cells_count, num_no_match_found

    def load_product_expiry_list(self):
        product_expiry_dict = {"product_names": list, "stock": list, "expiry_date": list}
        product_expiry_path = self.download_onedrive_file()
        product_expiry_workbook = xl.load_workbook(product_expiry_path)
        product_expiry_ws = product_expiry_workbook.worksheets[0]
        mr_prod_expiry = product_expiry_ws.max_row
        mc_prod_expiry = product_expiry_ws.max_column
        # column of stock destination
        product_expiry_dict['product_names'] = product_expiry_ws['B'][1:]
        # column of stock destination
        product_expiry_dict['stock'] = product_expiry_ws['D'][1:]
        # category data
        product_expiry_dict['expiry_date'] = product_expiry_ws['E'][1:]

        return product_expiry_dict, mr_prod_expiry, mc_prod_expiry

    def load_ks_file_stock_sync(self):
        self.filepath_kassen_system = self.get_kassen_system_latest_file()
        ks_workbook = self.csv_to_excel()
        kassen_system_data_dict, mr_s, mc_s = self.load_kassen_system_excel_file(ks_workbook)
        return kassen_system_data_dict, mr_s, mc_s

    def get_expiry_products(self, product_expiry_dict):
        # load expiry date not required products
        with open("./expiry/no_expiry_products_list.txt") as txt:
            list_expiry_date_not_req = txt.read().split("\n")
        # expiry logic
        for col_nam, expiry_date, stock_value in zip(product_expiry_dict['product_names'],
                                                     product_expiry_dict['expiry_date'],
                                                     product_expiry_dict['stock']):
            if not col_nam.value in list_expiry_date_not_req:
                if expiry_date.value:
                    if stock_value.value == 0:
                        self.stock_zero_expiry_date_exists.append(col_nam.value)
                    if stock_value.value > 20:
                        # one month
                        if (expiry_date.value.date() <= self.one_months) and (
                                expiry_date.value.date() >= datetime.now().date()):
                            self.twenty_one_month_expiry_list.append(col_nam.value)
                        # two months
                        elif (expiry_date.value.date() <= self.two_months) and (
                                expiry_date.value.date() >= datetime.now().date()):
                            self.twenty_two_month_expiry_list.append(col_nam.value)
                        # three months
                        elif (expiry_date.value.date() <= self.three_months) and (
                                expiry_date.value.date() >= datetime.now().date()):
                            self.twenty_three_month_expiry_list.append(col_nam.value)
                        # products already expired
                        elif expiry_date.value.date() <= datetime.now().date():
                            self.products_already_expired_list.append(col_nam.value)
                            # print("products already expired {}".format(col_nam.value))
                        else:
                            continue
                    if 11 <= stock_value.value <= 20:
                        # one month
                        if (expiry_date.value.date() <= self.one_months) and (
                                expiry_date.value.date() >= datetime.now().date()):
                            self.ten_one_month_expiry_list.append(col_nam.value)
                        # two months
                        elif (expiry_date.value.date() <= self.two_months) and (
                                expiry_date.value.date() >= datetime.now().date()):
                            self.ten_two_month_expiry_list.append(col_nam.value)
                        # three months
                        elif (expiry_date.value.date() <= self.three_months) and (
                                expiry_date.value.date() >= datetime.now().date()):
                            self.ten_three_month_expiry_list.append(col_nam.value)
                        # products already expired
                        elif expiry_date.value.date() <= datetime.now().date():
                            self.products_already_expired_list.append(col_nam.value)
                            # print("products already expired {}".format(col_nam.value))
                        else:
                            continue
                    if 1 <= stock_value.value <= 10:
                        # one month
                        if (expiry_date.value.date() <= self.one_months) and (
                                expiry_date.value.date() >= datetime.now().date()):
                            self.below_ten_one_month_expiry_list.append(col_nam.value)
                        # two months
                        elif (expiry_date.value.date() <= self.two_months) and (
                                expiry_date.value.date() >= datetime.now().date()):
                            self.below_ten_two_month_expiry_list.append(col_nam.value)
                        # three months
                        elif (expiry_date.value.date() <= self.three_months) and (
                                expiry_date.value.date() >= datetime.now().date()):
                            self.below_ten_three_month_expiry_list.append(col_nam.value)
                        # products already expired
                        elif expiry_date.value.date() <= datetime.now().date():
                            self.products_already_expired_list.append(col_nam.value)
                            # print("products already expired {}".format(col_nam.value))
                        else:
                            continue

                else:
                    if stock_value.value == 0:
                        self.no_expiry_date_products.append(col_nam.value)
                    else:
                        self.stock_exist_expiry_date_not_exists.append(col_nam.value)

    def notify_expiry_list(self, mr_prod_expiry, mr_s, num_match_products, num_no_match_products):
        subject = '[Notification] lotus-grocery.eu - Expiry Products in next 3 months ' + datetime.now().strftime(
            "%d/%m/%Y %H:%M:%S")
        self._text = self.text.format(
            stock201month=tabulate(zip(range(1, len(self.twenty_one_month_expiry_list) + 1), self.twenty_one_month_expiry_list),
                           headers=["Stock > 20 and 1 month\nno", "\nArticle Name"], tablefmt="grid"),
            stock202month=tabulate(zip(range(1, len(self.twenty_two_month_expiry_list) + 1), self.twenty_two_month_expiry_list),
                           headers=["Stock > 20 and 2 months\nno", "\nArticle Name"], tablefmt="grid"),
            stock203month=tabulate(
                zip(range(1, len(self.twenty_three_month_expiry_list) + 1), self.twenty_three_month_expiry_list),
                headers=["Stock > 20 and 3 months\nno", "\nArticle Name"], tablefmt="grid"),
            stock111month=tabulate(
                zip(range(1, len(self.ten_one_month_expiry_list) + 1), self.ten_one_month_expiry_list),
                headers=["Stock - (11-20) and 1 month\nno", "\nArticle Name"], tablefmt="grid"),
            stock112month=tabulate(
                zip(range(1, len(self.ten_two_month_expiry_list) + 1), self.ten_two_month_expiry_list),
                headers=["Stock - (11-20) and 2 months\nno", "\nArticle Name"], tablefmt="grid"),
            stock113month=tabulate(
                zip(range(1, len(self.ten_three_month_expiry_list) + 1), self.ten_three_month_expiry_list),
                headers=["Stock - (11-20) and 3 months\nno", "\nArticle Name"], tablefmt="grid"),
            stock11month=tabulate(
                zip(range(1, len(self.below_ten_one_month_expiry_list) + 1), self.below_ten_one_month_expiry_list),
                headers=["Stock - (1-10) and 1 month\nno", "\nArticle Name"], tablefmt="grid"),
            stock12month=tabulate(
                zip(range(1, len(self.below_ten_two_month_expiry_list) + 1), self.below_ten_two_month_expiry_list),
                headers=["Stock - (1-10) and 2 months\nno", "\nArticle Name"], tablefmt="grid"),
            stock13month=tabulate(
                zip(range(1, len(self.below_ten_three_month_expiry_list) + 1), self.below_ten_three_month_expiry_list),
                headers=["Stock - (1-10) and 3 months\nno", "\nArticle Name"], tablefmt="grid"),
            expiredproducts=tabulate(
                zip(range(1, len(self.products_already_expired_list) + 1), self.products_already_expired_list),
                headers=["Products already Expired\nno", "\nArticle Name"], tablefmt="grid"),
            case1=tabulate(
                zip(range(1, len(self.stock_zero_expiry_date_exists) + 1), self.stock_zero_expiry_date_exists),
                headers=["Case-1:Stock = 0, expiry date !=0 \nno", "\nArticle Name"], tablefmt="grid"),
            case2=tabulate(
                zip(range(1, len(self.stock_exist_expiry_date_not_exists) + 1), self.stock_exist_expiry_date_not_exists),
                headers=["Case-2:Stock != 0 expiry date !=1\nno", "\nArticle Name"], tablefmt="grid"),
            stat=tabulate([['Total no of Rows/Products in product expiry list/website', mr_prod_expiry], ['Total no of Rows/Products in kassen_system_file', mr_s],
                           ['Number of Products Matched', num_match_products],
                           ['Number of Products are no matched', num_no_match_products]], headers=['Stat Name', 'Stat Value'],tablefmt="grid")
        )
        self._html = self.html.format(
            stock201month=tabulate(
                zip(range(1, len(self.twenty_one_month_expiry_list) + 1), self.twenty_one_month_expiry_list),
                headers=["Stock > 20 and 1 month\nno", "\nArticle Name"], tablefmt="html"),
            stock202month=tabulate(
                zip(range(1, len(self.twenty_two_month_expiry_list) + 1), self.twenty_two_month_expiry_list),
                headers=["Stock > 20 and 2 months\nno", "\nArticle Name"], tablefmt="html"),
            stock203month=tabulate(
                zip(range(1, len(self.twenty_three_month_expiry_list) + 1), self.twenty_three_month_expiry_list),
                headers=["Stock > 20 and 3 months\nno", "\nArticle Name"], tablefmt="html"),
            stock111month=tabulate(
                zip(range(1, len(self.ten_one_month_expiry_list) + 1), self.ten_one_month_expiry_list),
                headers=["Stock - (11-20) and 1 month\nno", "\nArticle Name"], tablefmt="html"),
            stock112month=tabulate(
                zip(range(1, len(self.ten_two_month_expiry_list) + 1), self.ten_two_month_expiry_list),
                headers=["Stock - (11-20) and 2 months\nno", "\nArticle Name"], tablefmt="html"),
            stock113month=tabulate(
                zip(range(1, len(self.ten_three_month_expiry_list) + 1), self.ten_three_month_expiry_list),
                headers=["Stock - (11-20) and 3 months\nno", "\nArticle Name"], tablefmt="html"),
            stock11month=tabulate(
                zip(range(1, len(self.below_ten_one_month_expiry_list) + 1), self.below_ten_one_month_expiry_list),
                headers=["Stock - (1-10) and 1 month\nno", "\nArticle Name"], tablefmt="html"),
            stock12month=tabulate(
                zip(range(1, len(self.below_ten_two_month_expiry_list) + 1), self.below_ten_two_month_expiry_list),
                headers=["Stock - (1-10) and 2 months\nno", "\nArticle Name"], tablefmt="html"),
            stock13month=tabulate(
                zip(range(1, len(self.below_ten_three_month_expiry_list) + 1), self.below_ten_three_month_expiry_list),
                headers=["Stock - (1-10) and 3 months\nno", "\nArticle Name"], tablefmt="html"),
            expiredproducts=tabulate(
                zip(range(1, len(self.products_already_expired_list) + 1), self.products_already_expired_list),
                headers=["Products already Expired\nno", "\nArticle Name"], tablefmt="html"),
            case1=tabulate(
                zip(range(1, len(self.stock_zero_expiry_date_exists) + 1), self.stock_zero_expiry_date_exists),
                headers=["Case-1:Stock = 0, expiry date !=0: \nno", "\nArticle Name"], tablefmt="html"),
            case2=tabulate(
                zip(range(1, len(self.stock_exist_expiry_date_not_exists) + 1),
                    self.stock_exist_expiry_date_not_exists),
                headers=["Case-2:Stock != 0 expiry date !=1 \nno", "\nArticle Name"],
                tablefmt="html"),
            stat=tabulate([['Total no of Rows/Products in product expiry list/website', mr_prod_expiry],
                           ['Total no of Rows/Products in kassen_system_file', mr_s],
                           ['Number of Products Matched', num_match_products],
                           ['Number of Products are no matched', num_no_match_products]],
                          headers=['Stat Name', 'Stat Value'], tablefmt="html")
        )
        # content = [message]
        # self.send_email(subject, content)
        self.send_email_using_smtp(self._text, self._html, subject, "stocks@lotus-grocery.eu", "info@lotus-grocery.eu")

    def process(self):
        product_expiry_dict, mr_prod_expiry, mc_prod_expiry = self.load_product_expiry_list()
        kassen_system_data_dict, mr_s, mc_s = self.load_ks_file_stock_sync()
        num_match_products, num_no_match_products = self.sync_stock_from_ks_to_expiry_list(product_expiry_dict,
                                                                                           kassen_system_data_dict)
        self.get_expiry_products(product_expiry_dict)
        self.notify_expiry_list(mr_prod_expiry, mr_s, num_match_products, num_no_match_products)
        print(self._text)


if __name__ == "__main__":
    kassen_system_excel_file_dir = r"/Users/muralidharpettela/Downloads/ks_dir"
    product_expiry_list_dir = r"/Users/muralidharpettela/Downloads/product_expiry"
    onedrive_link = "https://1drv.ms/x/s!Auk2yZWl9__ZguU7J96x9U_HkGBDAQ?e=FYcE7A"
    credentials_path = "./common/email_credentials.json"
    expiry_products = ExpiryProducts(kassen_system_excel_file_dir, product_expiry_list_dir, onedrive_link, credentials_path)
    expiry_products.process()
