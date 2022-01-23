import openpyxl as xl
from woocommerce import API
import json
import csv
import sys
from openpyxl.workbook import Workbook
##############
filepath_kassen_system = r"/Users/muralidharpettela/Downloads/BK_Artikeldaten_22012022.xlsx"
json_file_path = "products.json"
kassen_system_data_dict = {"product_names": list, "stock": list, "price": list, "sale_price": list, "tax_class": list}
products_list = list()
no_match_products_list = list()
no_match_products_txt = open("no_match_products.txt", "w+")
##############

wcapi = API(
    url="https://www.lotus-grocery.eu/",
    consumer_key="ck_a1a83db1a7931bc4c965bf0e3d281ac63bea7264",
    consumer_secret="cs_3fd453e8a22d1d3da2a1376c1d8906130f6de1e4",
    timeout=1000
)


def csv_to_excel(input_csv_file, delimiter=";"):
    if not ".csv" in input_csv_file:
        sys.stderr.write("Error: File does not have the ending \".csv\".\n")
        sys.exit(2)

    input_fh = open(input_csv_file, encoding="ISO-8859-1")
    workbook = Workbook()
    #sheet = workbook.create_sheet(0)
    sheet = workbook.active

    for row_index, row in enumerate(csv.reader(open(input_csv_file, encoding="ISO-8859-1"), delimiter=delimiter)):
        for col_index, col in enumerate(row):
            if row_index > 0:
                if col_index == 2 or col_index == 3 or col_index == 4 or col_index == 5:
                    sheet.cell(row=row_index + 1, column=col_index + 1).value = float(col.replace(",", "."))
                else:
                    sheet.cell(row=row_index + 1, column=col_index + 1).value = col
            else:
                sheet.cell(row=row_index + 1, column=col_index + 1).value = col

    workbook.save(open(input_csv_file.replace(".csv", ".xlsx"), "wb"))
    return input_csv_file.replace(".csv", ".xlsx")

# Source coming from shop
# Destination products in website sheet
# opening the source excel file
def load_kassen_system_excel_file(filename_kassen_system_path):
    wb1 = xl.load_workbook(filename_kassen_system_path)
    ws1 = wb1.worksheets[0]
    # calculate total number of rows and
    # columns in source excel file
    mr_s = ws1.max_row
    mc_s = ws1.max_column

    return ws1, mr_s, mc_s


def load_json_data_website_products(json_file_path):
    # load the json file
    # Opening JSON file
    f = open(json_file_path)
    # returns JSON object as
    # a dictionary
    data = json.load(f)
    return data


def assign_data_from_ks(ws1):
    # column of sort id source file
    kassen_system_data_dict["product_names"] = ws1['B'][1:]
    # column of stock Source file
    kassen_system_data_dict["stock"] = ws1['F'][1:]
    # column of stock Source file
    kassen_system_data_dict["price"] = ws1['C'][1:]
    # Kassen system sale price (EK)
    kassen_system_data_dict["sale_price"] = ws1['D'][1:]
    # tax class
    kassen_system_data_dict["tax_class"] = ws1['E'][1:]

    return kassen_system_data_dict


def calculate_weight(product_website, products_kassen_system_dict):
    # weight calculate
    if str(product_website['name']).rstrip().count("-") > 1:
        # index_of_ = str(product_website.value).rstrip().index("-")
        # new_product_name = str(product_website.value)[index_of_].replace("-", " ")
        pass
    else:
        if len(str(product_website['name']).rstrip().split("-")[1].split()) > 2:
            weight_grams_list = str(product_website['name']).rstrip().split("-")[1].split()[:-1]
        else:
            weight_grams_list = str(product_website['name']).rstrip().split("-")[1].split()
        try:
            if weight_grams_list[1] == "g":
                weight = float(weight_grams_list[0]) / 1000
                # weight
                products_kassen_system_dict['weight'] = str(weight).replace(",", ".")
            elif weight_grams_list[1] == "kg":
                weight = float(weight_grams_list[0])
                # weight
                products_kassen_system_dict['weight'] = str(weight).replace(",", ".")
            elif weight_grams_list[1] == "l":
                weight = float(weight_grams_list[0])
                # weight
                products_kassen_system_dict['weight'] = str(weight).replace(",", ".")
            elif weight_grams_list[1] == "ml":
                weight = float(weight_grams_list[0]) / 1000
                # weight
                products_kassen_system_dict['weight'] = str(weight).replace(",", ".")
            else:
                pass
        except:
            print(str(product_website['name']).rstrip())


def match_products_and_update(json_data_dict, kassen_system_data_dict):
    num_no_match_found = 0
    match_of_stock_cells_count = 0
    for i, product_website in enumerate(json_data_dict):
        products_kassen_system_dict = {"id": 0, "weight": "0", "stock_quantity": 0, "regular_price": 0,
                                       "sale_price": None,
                                       "tax_class": None, }
        for j, product_kassen_system in enumerate(kassen_system_data_dict["product_names"]):
            # check the sort id source and destination are same, if yes update the stock of destination with stock of source
            if str(product_website['name']).rstrip() == str(product_kassen_system.value).rstrip():
                products_kassen_system_dict['id'] = product_website["id"]
                # calculate weight
                calculate_weight(product_website, products_kassen_system_dict)
                # stock update
                products_kassen_system_dict['stock_quantity'] = kassen_system_data_dict["stock"][j].value
                # price update
                # products_kassen_system_dict['price'] = str(kassen_system_data_dict["price"][j].value)
                # sale price
                if kassen_system_data_dict["sale_price"][j].value != 0:
                    products_kassen_system_dict['sale_price'] = str(kassen_system_data_dict["price"][j].value)
                    products_kassen_system_dict['regular_price'] = str(kassen_system_data_dict["sale_price"][j].value)
                else:
                    # regular price update
                    products_kassen_system_dict['regular_price'] = str(kassen_system_data_dict["price"][j].value)
                # tax class update
                if kassen_system_data_dict["tax_class"][j].value == 7:
                    products_kassen_system_dict['tax_class'] = "Tax 7 Per"
                else:
                    products_kassen_system_dict['tax_class'] = "Tax 19 Per"

                # print(wcapi.put("products/" + str(product_website["id"]), products_kassen_system_dict).json())
                # wcapi.put("products/" + str(product_website["id"]), products_kassen_system_dict).json()
                products_list.append(products_kassen_system_dict)
                match_of_stock_cells_count = match_of_stock_cells_count + 1
                break

            if (j == len(kassen_system_data_dict["product_names"]) - 1):
                if str(product_website['name']) not in no_match_products_list:
                    no_match_products_list.append(product_website['name'])
                    no_match_products_txt.write(product_website['name'])
                    no_match_products_txt.write("\n")
                    num_no_match_found = num_no_match_found + 1

    return match_of_stock_cells_count, num_no_match_found


def chunks(l, n):
    """Yield successive n-sized chunks from l."""
    for i in range(0, len(l), n):
        yield l[i:i + n]


def main():
    import timeit

    start = timeit.default_timer()
    path_of_excel = csv_to_excel(filepath_kassen_system)
    ws1, mr_s, mc_s = load_kassen_system_excel_file(path_of_excel)
    json_data = load_json_data_website_products(json_file_path)
    kassen_system_data = assign_data_from_ks(ws1)
    match_of_stock_cells_count, num_no_match_found = match_products_and_update(json_data, kassen_system_data)
    # json_string = json.dumps({"update": products_list})
    MAX_API_BATCH_SIZE = 100
    for batch in chunks(products_list, MAX_API_BATCH_SIZE):
        print(len(batch))
        print(wcapi.put("products/batch", {"update": batch}).json())
    # with open("products_all.json", "w") as jsonfile:
    # jsonfile.write(json_string)
    # jsonfile.close()
    print("Total no of Rows/Products in Source file from Shop File:{}".format(mr_s))
    print("Total no of Rows/Products in Destination file in Website:{}".format(len(json_data)))
    print("Number of Products Matched:{}".format(match_of_stock_cells_count))
    print("Number of Products are no matched:{}".format(num_no_match_found))
    stop = timeit.default_timer()

    print('Time: ', stop - start)


if __name__ == "__main__":
    main()
