import openpyxl as xl
import base64
import yagmail
import glob
import os
import requests
import csv
import sys
from datetime import datetime, date
from datetime import date
from dateutil.relativedelta import relativedelta

# to sync the stock
kassen_system_excel_file_dir = r"/Users/muralidharpettela/Downloads/"
product_expiry_list_dir = r"/Users/muralidharpettela/Downloads/product_expiry"

one_months = date.today() + relativedelta(months=+1)
two_months = date.today() + relativedelta(months=+2)
three_months = date.today() + relativedelta(months=+3)

one_month_expiry_list = list()
two_month_expiry_list = list()
three_month_expiry_list = list()
products_already_expired_list = list()
no_expiry_date_products = list()
stock_zero_expiry_date_exists = list()
stock_exist_expiry_date_not_exists = list()
# new_products_workbook = xl.load_workbook(product_expiry_list_dir)
# new_products_ws = new_products_workbook.worksheets[0]
# column of sort id source file
# col_name = new_products_ws['B'][1:]
# category data
# col_category = new_products_ws['E'][1:]


def create_onedrive_directdownload(onedrive_link):
    data_bytes64 = base64.b64encode(bytes(onedrive_link, 'utf-8'))
    data_bytes64_String = data_bytes64.decode('utf-8').replace('/', '_').replace('+', '-').rstrip("=")
    resultUrl = f"https://api.onedrive.com/v1.0/shares/u!{data_bytes64_String}/root/content"
    return resultUrl


def send_email(subject, content):
    user = 'lotusgroceryingolstadt2@gmail.com'
    app_password = 'ofanobixbzpvcpqz'  # a token for gmail
    to = 'info@lotus-grocery.eu'  # To send a group of recipients, simply change ‘to’ to a list.

    subject = subject
    content = content

    with yagmail.SMTP(user, app_password) as yag:
        yag.send(to, subject, content)
        print('Sent email successfully')


def download_onedrive_file(onedrive_link):
    onedrive_direct_link = create_onedrive_directdownload(onedrive_link)
    r = requests.get(onedrive_direct_link)
    save_path = os.path.join(product_expiry_list_dir, "products_expiry_list.xlsx")
    with open(save_path, 'wb') as f:
        f.write(r.content)
    return save_path


def csv_to_excel(filename_kassen_system):
    if not ".csv" in filename_kassen_system:
        sys.stderr.write("Error: File does not have the ending \".csv\".\n")
        sys.exit(2)

    input_fh = open(filename_kassen_system, encoding="ISO-8859-1")
    workbook = xl.Workbook()
    # sheet = workbook.create_sheet(0)
    sheet = workbook.active

    for row_index, row in enumerate(
            csv.reader(open(filename_kassen_system, encoding="ISO-8859-1"), delimiter=";")):
        for col_index, col in enumerate(row):
            if row_index > 0:
                if col_index == 2 or col_index == 3 or col_index == 4 or col_index == 5:
                    sheet.cell(row=row_index + 1, column=col_index + 1).value = float(col.replace(",", "."))
                else:
                    sheet.cell(row=row_index + 1, column=col_index + 1).value = col
            else:
                sheet.cell(row=row_index + 1, column=col_index + 1).value = col

    # workbook.save(open(input_csv_file.replace(".csv", ".xlsx"), "wb"))
    return workbook


# kassen system latest file
list_of_files = glob.glob(kassen_system_excel_file_dir + "/*.csv")  # * means all if need specific format then *.csv
latest_file = max(list_of_files, key=os.path.getctime)
print(latest_file)

onedrive_link = "https://1drv.ms/x/s!Auk2yZWl9__ZguNE2gKSuLVF55_waA?e=a5rfd9"
product_expiry_path = download_onedrive_file(onedrive_link)
product_expiry_workbook = xl.load_workbook(product_expiry_path)
product_expiry_ws = product_expiry_workbook.worksheets[0]

# case 1: stock 0 but having expiry date, Check Stock and update expiry date or delete expiry dates
# sync the stock from ks excel list to product expiry list
kassen_system_workbook = csv_to_excel(latest_file)
ws1 = kassen_system_workbook.worksheets[0]
# calculate total number of rows and
# columns in source excel file
mr_s = ws1.max_row
mc_s = ws1.max_column

mr_d = product_expiry_ws.max_row
mc_d = product_expiry_ws.max_column
# column of sort id source file
product_names_kassen_system = ws1['B'][1:]
# column of stock Source file
stock_kassen_system = ws1['F'][1:]
# column of stock Source file
price_kassen_system = ws1['C'][1:]
# Kassen system sale price (EK)
sale_price_kassen_system = ws1['D'][1:]
# tax class
tax_class_kassen_system = ws1['E'][1:]

# column of stock destination
product_names_expiry_list = product_expiry_ws['B'][1:]
# column of stock destination
stock_expiry_list = product_expiry_ws['D'][1:]
# category data
expiry_date_expiry_list = product_expiry_ws['E'][1:]

match_of_stock_cells_count = 0
num_of_product_stock_changed = 0
num_of_product_price_changed = 0
num_of_tax_class_changed = 0
num_no_match_found = 0
num_of_sale_price_updates = 0

no_match_products_list = list()
no_match_products_txt = open("no_match_products.txt", "w+")
for i, product_website in enumerate(product_names_expiry_list):
    for j, product_kassen_system in enumerate(product_names_kassen_system):
        # check the sort id source and destination are same, if yes update the stock of destination with stock of source
        if str(product_website.value).rstrip() == str(product_kassen_system.value).rstrip():
            # stock update
            stock_expiry_list[i].value = stock_kassen_system[j].value
            match_of_stock_cells_count = match_of_stock_cells_count + 1
            break
        if (j == len(product_names_kassen_system) - 1):
            if str(product_website) not in no_match_products_list:
                no_match_products_list.append(product_website.value)
                no_match_products_txt.write(product_website.value)
                no_match_products_txt.write("\n")
                num_no_match_found = num_no_match_found + 1
print("Total no of Rows/Products in Source file from Shop File:{}".format(mr_s))
print("Total no of Rows/Products in Destination file in Website:{}".format(mr_d))
print("Number of Products Matched:{}".format(match_of_stock_cells_count))
print("Number of Products Stock Changed:{}".format(num_of_product_stock_changed))
print("Number of Products Price Changed:{}".format(num_of_product_price_changed))
print("Number of Products Tax Class Changed:{}".format(num_of_tax_class_changed))
print("Number of Products are no matched:{}".format(num_no_match_found))
print("Number of Products Sale Price Changed:{}".format(num_of_sale_price_updates))
# # copying the cell values from source
# # excel file to destination excel file
# for i in range(1, mr + 1):
#     for j in range(1, mc + 1):
#         # reading cell value from source excel file
#         c = ws1.cell(row=i, column=j)
#
#         # writing the read value to destination excel file
#         ws2.cell(row=i, column=j).value = c.value

# saving the destination excel file
product_expiry_workbook.save(str(product_expiry_path))
# expiry logic
for col_nam, expiry_date, stock_value in zip(product_names_expiry_list, expiry_date_expiry_list, stock_expiry_list):
    if expiry_date.value:
        if stock_value.value == 0:
            stock_zero_expiry_date_exists.append(col_nam.value)
        # one month
        if (expiry_date.value.date() <= one_months) and (expiry_date.value.date() >= datetime.now().date()):
            one_month_expiry_list.append(col_nam.value)
        # two months
        if (expiry_date.value.date() <= two_months) and (expiry_date.value.date() >= datetime.now().date()):
            two_month_expiry_list.append(col_nam.value)
        # three months
        if (expiry_date.value.date() <= three_months) and (expiry_date.value.date() >= datetime.now().date()):
            three_month_expiry_list.append(col_nam.value)
        # products already expired
        if expiry_date.value.date() <= datetime.now().date():
            products_already_expired_list.append(col_nam.value)
            # print("products already expired {}".format(col_nam.value))
    else:
        if stock_value.value == 0:
            no_expiry_date_products.append(col_nam.value)
        else:
            stock_exist_expiry_date_not_exists.append(col_nam.value)



message = "This is an automated mail, notification of the expiry products within 3 months. \n " \
          "Products expiring in 1 month:\n{}\n{}\n{}\n" \
          "Products expiring in 2 months:\n{}\n{}\n{}\n" \
          "Products expiring in 3 month:\n{}\n{}\n{}\n" \
          "Products already expired:\n{}\n{}\n{}\n"\
          "Case-1: Stock value 0 but the expiry date exists. Please recheck the products stock and update expiry dates\n{}\n{}\n{}"\
          "Case-2: Stock value is not zero but the expiry date doesnot exists. Please recheck the products stock and update expiry dates\n{}\n{}\n{}".format("==================================",
                                                           "\n".join(one_month_expiry_list),
                                                           "==================================",
                                                           "==================================",
                                                           "\n".join(two_month_expiry_list),
                                                           "==================================",
                                                           "==================================",
                                                           "\n".join(three_month_expiry_list),
                                                           "==================================",
                                                           "==================================",
                                                           "\n".join(products_already_expired_list),
                                                           "==================================", "==================================",
                                                     "\n".join(stock_zero_expiry_date_exists),
                                                     "==================================","==================================",
                                                     "\n".join(stock_exist_expiry_date_not_exists),
                                                     "==================================")
subject = '[Notification] lotus-grocery.eu - Expiry Products in next 3 months ' + datetime.now().strftime(
    "%d/%m/%Y %H:%M:%S")
content = [message]
send_email(subject, content)
print("Products expiring in 1 month:\n{}\n{}\n{}".format("==================================",
                                                         "\n".join(one_month_expiry_list),
                                                         "=================================="))
print("Products expiring in 2 months:\n{}\n{}\n{}".format("==================================",
                                                          "\n".join(two_month_expiry_list),
                                                          "=================================="))
print("Products expiring in 3 month:\n{}\n{}\n{}".format("==================================",
                                                         "\n".join(three_month_expiry_list),
                                                         "=================================="))
print("Products already expired:\n{}\n{}\n{}".format("==================================",
                                                     "\n".join(products_already_expired_list),
                                                     "=================================="))
print("Case-1: Stock value 0 but the expiry date exists. Please recheck the products stock and update expiry dates\n{}\n{}\n{}".format("==================================",
                                                     "\n".join(stock_zero_expiry_date_exists),
                                                     "=================================="))
print("Case-2: Stock value is not zero but the expiry date doesnot exists. Please recheck the products stock and update expiry dates\n{}\n{}\n{}".format("==================================",
                                                     "\n".join(stock_exist_expiry_date_not_exists),
                                                     "=================================="))
